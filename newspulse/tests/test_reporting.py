"""Share of voice and the Excel deliverable (newspulse.reporting)."""

from __future__ import annotations

import datetime as dt
import io

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from newspulse.models import Analysis, Article, Base, Category, Client
from newspulse.reporting import client_workbook, share_of_voice


@pytest.fixture
def session():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, expire_on_commit=False)()


def _add(session, client, title, *, days_ago=1, relevance=5, alert=False, source="FAZ"):
    when = dt.datetime.now(dt.UTC) - dt.timedelta(days=days_ago)
    art = Article(
        title=title, url=f"https://ex.de/{title}", source=source, published_at=when,
        fetched_at=when, summary_text="s", language="de", title_hash=title[:8],
        author="Anna Muster",
    )
    session.add(art)
    session.flush()
    session.add(
        Analysis(
            article_id=art.id, client_id=client.id, summary="Zusammenfassung.",
            category=Category.FINANZEN, relevance_score=relevance,
            importance_score=7, is_alert=alert,
        )
    )


@pytest.fixture
def portfolio(session):
    """One mandate benchmarked against one rival, plus an unrelated mandate that
    must never appear in the first one's comparison."""
    mandate = Client(name="Alpha AG")
    rival = Client(name="Beta AG", is_competitor=True)
    unrelated = Client(name="Gamma AG")
    session.add_all([mandate, rival, unrelated])
    session.flush()
    mandate.competitors.append(rival)
    _add(session, mandate, "A1", alert=True)
    _add(session, mandate, "A2")
    _add(session, mandate, "A3")
    _add(session, rival, "B1")
    _add(session, unrelated, "G1")
    _add(session, unrelated, "G2")
    session.commit()
    return mandate, rival, unrelated


def test_share_of_voice_covers_the_client_and_its_own_competitors(session, portfolio):
    mandate, rival, _ = portfolio
    by_name = {v.name: v for v in share_of_voice(session, mandate, days=30)}
    assert set(by_name) == {"Alpha AG", "Beta AG"}
    assert by_name["Alpha AG"].mentions == 3
    assert by_name["Beta AG"].mentions == 1
    assert by_name["Alpha AG"].alerts == 1


def test_more_than_one_alert_is_counted_as_more_than_one(session):
    """Regression: the alert column could never read higher than 1.

    ``func.sum`` over a Boolean column inherits that column's result processor,
    so SQLAlchemy handed a total of 40 back as ``True`` and ``int(True)`` is 1.
    Every fixture here happened to have exactly one alert, which is precisely the
    value the bug produces — so the tests agreed with it. Measured on the dev
    database, a client with 40 alerts reported 1, on screen and in the client
    report both.
    """
    mandate = Client(name="Viel Alarm AG")
    session.add(mandate)
    session.flush()
    for i in range(4):
        _add(session, mandate, f"X{i}", alert=True)
    _add(session, mandate, "ruhig")
    session.commit()

    voice = {v.name: v for v in share_of_voice(session, mandate, days=30)}

    assert voice["Viel Alarm AG"].mentions == 5
    assert voice["Viel Alarm AG"].alerts == 4


def test_an_unrelated_mandate_never_enters_the_comparison(session, portfolio):
    """Share of voice is a statement about a market. "Zalando vs Siemens" is not
    a fact about anything, so a client outside the set must not appear."""
    mandate, _, unrelated = portfolio
    names = {v.name for v in share_of_voice(session, mandate, days=30)}
    assert unrelated.name not in names


def test_the_subject_is_flagged_relative_to_the_comparison_not_globally(session, portfolio):
    """A mandate can be someone else's benchmark, so is_competitor here means
    "not the client this comparison is about"."""
    mandate, rival, _ = portfolio
    by_name = {v.name: v for v in share_of_voice(session, mandate, days=30)}
    assert by_name["Alpha AG"].is_competitor is False
    assert by_name["Beta AG"].is_competitor is True


def test_shares_sum_to_one_within_the_comparison(session, portfolio):
    mandate, _, _ = portfolio
    voice = share_of_voice(session, mandate, days=30)
    assert sum(v.share for v in voice) == pytest.approx(1.0)
    assert {v.name: round(v.share, 2) for v in voice}["Alpha AG"] == 0.75


def test_a_competitor_with_no_coverage_still_appears_at_zero(session, portfolio):
    """Dropping the row would make the comparison look like it was never run."""
    mandate, _, _ = portfolio
    silent = Client(name="Delta AG", is_competitor=True)
    session.add(silent)
    session.flush()
    mandate.competitors.append(silent)
    session.commit()
    by_name = {v.name: v for v in share_of_voice(session, mandate, days=30)}
    assert by_name["Delta AG"].mentions == 0
    assert by_name["Delta AG"].share == 0.0


def test_a_client_with_no_competitors_is_a_comparison_of_one(session, portfolio):
    _, _, unrelated = portfolio
    voice = share_of_voice(session, unrelated, days=30)
    assert [v.name for v in voice] == ["Gamma AG"]
    assert voice[0].share == pytest.approx(1.0)


def test_a_month_with_no_coverage_at_all_is_zero_not_a_division_error(session):
    quiet = Client(name="Leer AG")
    session.add(quiet)
    session.commit()
    voice = share_of_voice(session, quiet, days=30)
    assert voice[0].mentions == 0
    assert voice[0].share == 0.0


def test_share_of_voice_respects_the_window(session, portfolio):
    mandate, _, _ = portfolio
    _add(session, mandate, "OLD", days_ago=200)
    session.commit()
    by_name = {v.name: v for v in share_of_voice(session, mandate, days=30)}
    assert by_name["Alpha AG"].mentions == 3


def test_irrelevant_analyses_never_reach_a_client_facing_number(session, portfolio):
    mandate, _, _ = portfolio
    _add(session, mandate, "NOISE", relevance=0)
    session.commit()
    by_name = {v.name: v for v in share_of_voice(session, mandate, days=30)}
    assert by_name["Alpha AG"].mentions == 3


def test_a_company_cannot_be_its_own_competitor(session):
    """Enforced in the schema so a self-link can never double-count a client."""
    from sqlalchemy.exc import IntegrityError

    from newspulse.models import client_competitors

    c = Client(name="Solo AG")
    session.add(c)
    session.commit()
    with pytest.raises(IntegrityError):
        session.execute(
            client_competitors.insert().values(client_id=c.id, competitor_id=c.id)
        )
        session.commit()
    session.rollback()


def test_workbook_leads_with_the_answers_then_the_evidence(session, portfolio):
    """The three questions a client actually asks — how are we doing against the
    competition, where should we be and are not, what should we say — used to
    live on three screens and leave the tool only as screenshots. They are the
    first sheets now, ahead of the raw list they are drawn from."""
    mandate, _, _ = portfolio
    book = load_workbook(io.BytesIO(client_workbook(session, mandate, days=30)))
    assert book.sheetnames == [
        "Überblick",
        "Share of Voice",
        "Pitch-Lücken",
        "Impulse",
        "Berichterstattung",
        "Nach Kategorie",
        "Nach Publisher",
    ]
    sheet = book["Berichterstattung"]
    headers = [cell.value for cell in sheet[1]]
    assert "Schlagzeile" in headers and "Autor" in headers and "Link" in headers
    assert sheet.max_row == 4  # header + three articles


def test_the_workbooks_share_of_voice_matches_the_dashboards(session, portfolio):
    """A number in the meeting and the same number on screen must not disagree."""
    mandate, _, _ = portfolio
    on_screen = {v.name: v.mentions for v in share_of_voice(session, mandate, days=30)}

    book = load_workbook(io.BytesIO(client_workbook(session, mandate, days=30)))
    sheet = book["Share of Voice"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    in_document = {row[0]: row[2] for row in rows if row[0]}

    assert in_document == on_screen


def test_the_overview_sheet_states_the_period_it_covers(session, portfolio):
    """A document that leaves the tool has to carry its own window: a reader in
    a meeting cannot ask the dashboard which month this was."""
    mandate, _, _ = portfolio
    book = load_workbook(io.BytesIO(client_workbook(session, mandate, days=30)))
    values = {
        row[0]: row[1]
        for row in book["Überblick"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }

    assert values["Mandant"] == mandate.name
    assert values["Zeitraum"] == "letzte 30 Tage"
    assert values["Meldungen"] == 3


def test_workbook_for_an_empty_month_is_still_a_readable_sheet(session):
    """An empty month must produce a correctly-shaped file, not a blank one the
    recipient cannot open."""
    c = Client(name="Leer AG")
    session.add(c)
    session.commit()
    book = load_workbook(io.BytesIO(client_workbook(session, c, days=30)))
    assert [cell.value for cell in book["Berichterstattung"][1]][0] == "Datum"


# --- Coverage map --------------------------------------------------------------


def test_coverage_map_finds_outlets_that_cover_rivals_but_never_the_client(session):
    """This is the pitch list: a publication with a standing relationship to the
    competition and none with us."""
    from newspulse.coverage_map import build

    mandate = Client(name="Alpha AG")
    rival = Client(name="Beta AG", is_competitor=True)
    session.add_all([mandate, rival])
    session.flush()
    mandate.competitors.append(rival)
    _add(session, mandate, "A1", source="Handelsblatt")
    _add(session, rival, "B1", source="Handelsblatt")
    _add(session, rival, "B2", source="InStyle")
    _add(session, rival, "B3", source="InStyle")
    session.commit()

    gaps = {row.source for row in build(session, mandate, days=90).gaps}
    assert gaps == {"InStyle"}          # rivals only
    assert "Handelsblatt" not in gaps   # already covers us


def test_a_single_rival_mention_is_not_a_pitch_gap(session):
    """A one-off mention is not a relationship; pitching at it wastes the contact."""
    from newspulse.coverage_map import build

    mandate = Client(name="Alpha AG")
    rival = Client(name="Beta AG", is_competitor=True)
    session.add_all([mandate, rival])
    session.flush()
    mandate.competitors.append(rival)
    _add(session, rival, "B1", source="Einmalblatt")
    session.commit()

    assert build(session, mandate, days=90).gaps == ()


def test_max_cell_is_never_zero_so_the_view_can_divide_by_it(session):
    from newspulse.coverage_map import build

    c = Client(name="Leer AG")
    session.add(c)
    session.commit()
    assert build(session, c, days=90).max_cell == 1
